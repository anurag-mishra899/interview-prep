#!/usr/bin/env python3
"""
Investment Analysis: SIP vs Loan+Lumpsum
Generates detailed Excel workbook with calculations
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

# Create workbook
wb = openpyxl.Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ============================================================================
# SHEET 1: SUMMARY & COMPARISON
# ============================================================================
ws_summary = wb.create_sheet("Summary & Comparison", 0)

# Headers
ws_summary['A1'] = "INVESTMENT ANALYSIS: SIP vs LOAN + LUMPSUM"
ws_summary['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_summary['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_summary.merge_cells('A1:F1')

# Scenario 1 Details
row = 3
ws_summary[f'A{row}'] = "SCENARIO 1: MONTHLY SIP"
ws_summary[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
ws_summary[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws_summary.merge_cells(f'A{row}:C{row}')

row += 1
ws_summary[f'A{row}'] = "Monthly SIP Amount"
ws_summary[f'B{row}'] = 40000
ws_summary[f'B{row}'].number_format = '₹#,##0'

row += 1
ws_summary[f'A{row}'] = "Annual Interest Rate"
ws_summary[f'B{row}'] = 0.12
ws_summary[f'B{row}'].number_format = '0.00%'

row += 1
ws_summary[f'A{row}'] = "Investment Period (Years)"
ws_summary[f'B{row}'] = 35

row += 1
ws_summary[f'A{row}'] = "Total Investment"
ws_summary[f'B{row}'] = "=B4*12*B6"
ws_summary[f'B{row}'].number_format = '₹#,##0'

row += 1
ws_summary[f'A{row}'] = "Future Value"
# FV = PMT × [((1 + r)^n - 1) / r] × (1 + r)
# where r = monthly rate, n = total months
ws_summary[f'B{row}'] = "=B4*((((1+B5/12)^(B6*12)-1)/(B5/12))*(1+B5/12))"
ws_summary[f'B{row}'].number_format = '₹#,##0'
ws_summary[f'B{row}'].font = Font(bold=True, color="006100")

row += 1
ws_summary[f'A{row}'] = "Net Gain"
ws_summary[f'B{row}'] = "=B9-B7"
ws_summary[f'B{row}'].number_format = '₹#,##0'

row += 1
ws_summary[f'A{row}'] = "ROI %"
ws_summary[f'B{row}'] = "=(B10/B7)*100"
ws_summary[f'B{row}'].number_format = '0.00"%"'

# Scenario 2 Details
row = 3
ws_summary[f'D{row}'] = "SCENARIO 2: LOAN + LUMPSUM"
ws_summary[f'D{row}'].font = Font(size=14, bold=True, color="FFFFFF")
ws_summary[f'D{row}'].fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
ws_summary.merge_cells(f'D{row}:F{row}')

row = 4
ws_summary[f'D{row}'] = "Loan Amount"
ws_summary[f'E{row}'] = 2500000
ws_summary[f'E{row}'].number_format = '₹#,##0'

row += 1
ws_summary[f'D{row}'] = "Loan Interest (Annual)"
ws_summary[f'E{row}'] = 0.10
ws_summary[f'E{row}'].number_format = '0.00%'

row += 1
ws_summary[f'D{row}'] = "Loan Tenure (Years)"
ws_summary[f'E{row}'] = 6

row += 1
ws_summary[f'D{row}'] = "Monthly EMI"
# EMI = P × r × (1+r)^n / ((1+r)^n - 1)
ws_summary[f'E{row}'] = "=E4*(E5/12)*(1+E5/12)^(E6*12)/((1+E5/12)^(E6*12)-1)"
ws_summary[f'E{row}'].number_format = '₹#,##0.00'

row += 1
ws_summary[f'D{row}'] = "Total EMI Paid (6 years)"
ws_summary[f'E{row}'] = "=E7*E6*12"
ws_summary[f'E{row}'].number_format = '₹#,##0'

row += 1
ws_summary[f'D{row}'] = "Investment Rate (Annual)"
ws_summary[f'E{row}'] = 0.12
ws_summary[f'E{row}'].number_format = '0.00%'

row += 1
ws_summary[f'D{row}'] = "Investment Period (Years)"
ws_summary[f'E{row}'] = 35

row += 1
ws_summary[f'D{row}'] = "Future Value"
ws_summary[f'E{row}'] = "=E4*(1+E9)^E10"
ws_summary[f'E{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'].font = Font(bold=True, color="C65911")

row += 1
ws_summary[f'D{row}'] = "Net Gain (after loan)"
ws_summary[f'E{row}'] = "=E11-E8"
ws_summary[f'E{row}'].number_format = '₹#,##0'

row += 1
ws_summary[f'D{row}'] = "ROI %"
ws_summary[f'E{row}'] = "=(E12/E8)*100"
ws_summary[f'E{row}'].number_format = '0.00"%"'

# Comparison Table
row = 15
ws_summary[f'A{row}'] = "COMPARATIVE ANALYSIS"
ws_summary[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
ws_summary[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_summary.merge_cells(f'A{row}:F{row}')

row += 1
headers = ["Metric", "Scenario 1 (SIP)", "Scenario 2 (Loan)", "Difference", "Winner"]
for col, header in enumerate(headers, 1):
    cell = ws_summary.cell(row, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

row += 1
ws_summary[f'A{row}'] = "Total Cash Outflow"
ws_summary[f'B{row}'] = "=B7"
ws_summary[f'B{row}'].number_format = '₹#,##0'
ws_summary[f'C{row}'] = "=E8"
ws_summary[f'C{row}'].number_format = '₹#,##0'
ws_summary[f'D{row}'] = "=B17-C17"
ws_summary[f'D{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'] = "=IF(C17<B17,\"Scenario 2\",\"Scenario 1\")"

row += 1
ws_summary[f'A{row}'] = "Future Value"
ws_summary[f'B{row}'] = "=B9"
ws_summary[f'B{row}'].number_format = '₹#,##0'
ws_summary[f'C{row}'] = "=E11"
ws_summary[f'C{row}'].number_format = '₹#,##0'
ws_summary[f'D{row}'] = "=B18-C18"
ws_summary[f'D{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'] = "=IF(B18>C18,\"Scenario 1\",\"Scenario 2\")"

row += 1
ws_summary[f'A{row}'] = "Net Gain"
ws_summary[f'B{row}'] = "=B10"
ws_summary[f'B{row}'].number_format = '₹#,##0'
ws_summary[f'C{row}'] = "=E12"
ws_summary[f'C{row}'].number_format = '₹#,##0'
ws_summary[f'D{row}'] = "=B19-C19"
ws_summary[f'D{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'] = "=IF(B19>C19,\"Scenario 1\",\"Scenario 2\")"

row += 1
ws_summary[f'A{row}'] = "ROI %"
ws_summary[f'B{row}'] = "=B11"
ws_summary[f'B{row}'].number_format = '0.00"%"'
ws_summary[f'C{row}'] = "=E13"
ws_summary[f'C{row}'].number_format = '0.00"%"'
ws_summary[f'D{row}'] = "=B20-C20"
ws_summary[f'D{row}'].number_format = '0.00"%"'
ws_summary[f'E{row}'] = "=IF(B20>C20,\"Scenario 1\",\"Scenario 2\")"

# Winner Box
row = 22
ws_summary[f'A{row}'] = "FINAL VERDICT"
ws_summary[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
ws_summary[f'A{row}'].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
ws_summary.merge_cells(f'A{row}:F{row}')

row += 1
ws_summary[f'A{row}'] = "WINNER:"
ws_summary[f'A{row}'].font = Font(size=12, bold=True)
ws_summary[f'B{row}'] = "SCENARIO 1 - MONTHLY SIP"
ws_summary[f'B{row}'].font = Font(size=12, bold=True, color="00B050")
ws_summary.merge_cells(f'B{row}:F{row}')

row += 1
ws_summary[f'A{row}'] = "Reason:"
ws_summary[f'B{row}'] = "SIP delivers significantly higher returns with lower risk, lower monthly commitment, and benefits from rupee cost averaging"
ws_summary.merge_cells(f'B{row}:F{row}')

# Column widths
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 18
ws_summary.column_dimensions['C'].width = 18
ws_summary.column_dimensions['D'].width = 25
ws_summary.column_dimensions['E'].width = 18
ws_summary.column_dimensions['F'].width = 15

# ============================================================================
# SHEET 2: YEAR-BY-YEAR SIP
# ============================================================================
ws_sip = wb.create_sheet("SIP Year-by-Year")

# Header
ws_sip['A1'] = "SCENARIO 1: SIP YEAR-BY-YEAR BREAKDOWN"
ws_sip['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws_sip['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws_sip.merge_cells('A1:E1')

# Column headers
headers = ["Year", "Total Invested", "Future Value", "Gain", "ROI %"]
for col, header in enumerate(headers, 1):
    cell = ws_sip.cell(2, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Parameters
monthly_sip = 40000
annual_rate = 0.12
monthly_rate = annual_rate / 12

# Calculate year by year
for year in range(1, 36):
    row = year + 2
    months = year * 12
    total_invested = monthly_sip * months

    # FV formula
    ws_sip[f'A{row}'] = year
    ws_sip[f'B{row}'] = total_invested
    ws_sip[f'B{row}'].number_format = '₹#,##0'

    # FV = PMT × [((1 + r)^n - 1) / r] × (1 + r)
    ws_sip[f'C{row}'] = f"=40000*((((1+0.01)^{months}-1)/0.01)*(1+0.01))"
    ws_sip[f'C{row}'].number_format = '₹#,##0'

    ws_sip[f'D{row}'] = f"=C{row}-B{row}"
    ws_sip[f'D{row}'].number_format = '₹#,##0'

    ws_sip[f'E{row}'] = f"=(D{row}/B{row})*100"
    ws_sip[f'E{row}'].number_format = '0.00"%"'

# Column widths
ws_sip.column_dimensions['A'].width = 8
ws_sip.column_dimensions['B'].width = 18
ws_sip.column_dimensions['C'].width = 18
ws_sip.column_dimensions['D'].width = 18
ws_sip.column_dimensions['E'].width = 12

# ============================================================================
# SHEET 3: YEAR-BY-YEAR LOAN + LUMPSUM
# ============================================================================
ws_loan = wb.create_sheet("Loan+Lumpsum Year-by-Year")

# Header
ws_loan['A1'] = "SCENARIO 2: LOAN + LUMPSUM YEAR-BY-YEAR BREAKDOWN"
ws_loan['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws_loan['A1'].fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
ws_loan.merge_cells('A1:F1')

# Column headers
headers = ["Year", "Loan Outstanding", "EMI Paid (Annual)", "Cumulative Paid", "Investment Value", "Net Worth"]
for col, header in enumerate(headers, 1):
    cell = ws_loan.cell(2, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Loan parameters
loan_amount = 2500000
loan_rate = 0.10
loan_rate_monthly = loan_rate / 12
loan_months = 72
investment_rate = 0.12

# Calculate EMI
emi = loan_amount * loan_rate_monthly * (1 + loan_rate_monthly)**loan_months / ((1 + loan_rate_monthly)**loan_months - 1)

# Calculate year by year
outstanding = loan_amount
for year in range(1, 36):
    row = year + 2

    ws_loan[f'A{row}'] = year

    # Loan calculations for first 6 years
    if year <= 6:
        annual_emi = emi * 12

        # Calculate outstanding at end of year
        for month in range(12):
            interest = outstanding * loan_rate_monthly
            principal = emi - interest
            outstanding -= principal

        ws_loan[f'B{row}'] = max(0, outstanding)
        ws_loan[f'B{row}'].number_format = '₹#,##0'

        ws_loan[f'C{row}'] = annual_emi
        ws_loan[f'C{row}'].number_format = '₹#,##0'

        ws_loan[f'D{row}'] = f"=C{row}" if year == 1 else f"=D{row-1}+C{row}"
        ws_loan[f'D{row}'].number_format = '₹#,##0'
    else:
        ws_loan[f'B{row}'] = 0
        ws_loan[f'B{row}'].number_format = '₹#,##0'

        ws_loan[f'C{row}'] = 0
        ws_loan[f'C{row}'].number_format = '₹#,##0'

        ws_loan[f'D{row}'] = f"=D{row-1}"
        ws_loan[f'D{row}'].number_format = '₹#,##0'

    # Investment value
    investment_value = loan_amount * (1 + investment_rate)**year
    ws_loan[f'E{row}'] = investment_value
    ws_loan[f'E{row}'].number_format = '₹#,##0'

    # Net worth
    ws_loan[f'F{row}'] = f"=E{row}-D{row}"
    ws_loan[f'F{row}'].number_format = '₹#,##0'

# Column widths
ws_loan.column_dimensions['A'].width = 8
ws_loan.column_dimensions['B'].width = 18
ws_loan.column_dimensions['C'].width = 18
ws_loan.column_dimensions['D'].width = 18
ws_loan.column_dimensions['E'].width = 18
ws_loan.column_dimensions['F'].width = 18

# ============================================================================
# SHEET 4: FORMULAS & METHODOLOGY
# ============================================================================
ws_formulas = wb.create_sheet("Formulas & Methodology")

ws_formulas['A1'] = "FORMULAS AND CALCULATION METHODOLOGY"
ws_formulas['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws_formulas['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_formulas.merge_cells('A1:D1')

row = 3
ws_formulas[f'A{row}'] = "1. SIP FUTURE VALUE FORMULA"
ws_formulas[f'A{row}'].font = Font(bold=True, size=12)
row += 1
ws_formulas[f'A{row}'] = "FV = PMT × [((1 + r)^n - 1) / r] × (1 + r)"
row += 1
ws_formulas[f'A{row}'] = "Where:"
row += 1
ws_formulas[f'A{row}'] = "  PMT = Monthly SIP amount (₹40,000)"
row += 1
ws_formulas[f'A{row}'] = "  r = Monthly interest rate (12%/12 = 1%)"
row += 1
ws_formulas[f'A{row}'] = "  n = Total number of months (35 years × 12 = 420 months)"

row += 2
ws_formulas[f'A{row}'] = "2. LUMPSUM FUTURE VALUE FORMULA"
ws_formulas[f'A{row}'].font = Font(bold=True, size=12)
row += 1
ws_formulas[f'A{row}'] = "FV = P × (1 + r)^n"
row += 1
ws_formulas[f'A{row}'] = "Where:"
row += 1
ws_formulas[f'A{row}'] = "  P = Principal amount (₹25,00,000)"
row += 1
ws_formulas[f'A{row}'] = "  r = Annual interest rate (12%)"
row += 1
ws_formulas[f'A{row}'] = "  n = Number of years (35 years)"

row += 2
ws_formulas[f'A{row}'] = "3. EMI CALCULATION FORMULA"
ws_formulas[f'A{row}'].font = Font(bold=True, size=12)
row += 1
ws_formulas[f'A{row}'] = "EMI = P × r × (1+r)^n / [(1+r)^n - 1]"
row += 1
ws_formulas[f'A{row}'] = "Where:"
row += 1
ws_formulas[f'A{row}'] = "  P = Loan principal (₹25,00,000)"
row += 1
ws_formulas[f'A{row}'] = "  r = Monthly interest rate (10%/12 = 0.833%)"
row += 1
ws_formulas[f'A{row}'] = "  n = Loan tenure in months (6 years × 12 = 72 months)"

row += 2
ws_formulas[f'A{row}'] = "4. ASSUMPTIONS"
ws_formulas[f'A{row}'].font = Font(bold=True, size=12)
row += 1
ws_formulas[f'A{row}'] = "  • SIP returns: 12% per annum (compounded monthly)"
row += 1
ws_formulas[f'A{row}'] = "  • Lumpsum returns: 12% per annum (compounded annually)"
row += 1
ws_formulas[f'A{row}'] = "  • Loan interest: 10% per annum"
row += 1
ws_formulas[f'A{row}'] = "  • No tax implications considered"
row += 1
ws_formulas[f'A{row}'] = "  • No inflation adjustments"
row += 1
ws_formulas[f'A{row}'] = "  • Regular payments without defaults"

ws_formulas.column_dimensions['A'].width = 80

# Save workbook
wb.save('/Users/anuragmishra/Code/interview-prep/Investment_Analysis_SIP_vs_Loan.xlsx')
print("Excel file created successfully: Investment_Analysis_SIP_vs_Loan.xlsx")
print("\nKey Results:")
print("=" * 60)

# Calculate final values
sip_months = 35 * 12
sip_monthly_rate = 0.12 / 12
sip_fv = 40000 * (((1 + sip_monthly_rate)**sip_months - 1) / sip_monthly_rate) * (1 + sip_monthly_rate)
sip_invested = 40000 * sip_months
sip_gain = sip_fv - sip_invested

loan_monthly_rate = 0.10 / 12
emi = 2500000 * loan_monthly_rate * (1 + loan_monthly_rate)**72 / ((1 + loan_monthly_rate)**72 - 1)
total_emi = emi * 72
lumpsum_fv = 2500000 * (1.12**35)
net_gain_loan = lumpsum_fv - total_emi

print(f"\nSCENARIO 1 (SIP):")
print(f"  Total Investment: ₹{sip_invested:,.0f}")
print(f"  Future Value: ₹{sip_fv:,.0f}")
print(f"  Net Gain: ₹{sip_gain:,.0f}")
print(f"  ROI: {(sip_gain/sip_invested)*100:.2f}%")

print(f"\nSCENARIO 2 (Loan + Lumpsum):")
print(f"  Monthly EMI: ₹{emi:,.2f}")
print(f"  Total EMI Paid: ₹{total_emi:,.0f}")
print(f"  Future Value: ₹{lumpsum_fv:,.0f}")
print(f"  Net Gain: ₹{net_gain_loan:,.0f}")
print(f"  ROI: {(net_gain_loan/total_emi)*100:.2f}%")

print(f"\nWINNER: SCENARIO 1 (SIP)")
print(f"  Higher returns by: ₹{sip_gain - net_gain_loan:,.0f}")
print("=" * 60)
