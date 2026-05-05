#!/usr/bin/env python3
"""
6-Year Investment Comparison: Loan + Lumpsum vs SIP
Both scenarios have SAME monthly cash outflow
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import math

# Create workbook
wb = openpyxl.Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ============================================================================
# CALCULATE KEY VALUES FIRST
# ============================================================================

# Loan parameters
loan_amount = 2500000
loan_rate_annual = 0.10
loan_rate_monthly = loan_rate_annual / 12
loan_months = 72

# Calculate EMI
emi = loan_amount * loan_rate_monthly * (1 + loan_rate_monthly)**loan_months / ((1 + loan_rate_monthly)**loan_months - 1)

# Investment parameters
investment_rate_annual = 0.12
investment_rate_monthly = investment_rate_annual / 12

# SCENARIO 1: Loan + Lumpsum
# Lumpsum investment grows for 6 years
lumpsum_value_6y = loan_amount * (1 + investment_rate_annual)**6
total_emi_paid = emi * loan_months
net_value_scenario1 = lumpsum_value_6y - total_emi_paid

# SCENARIO 2: SIP with same EMI amount
# SIP of EMI amount for 6 years
sip_value_6y = emi * (((1 + investment_rate_monthly)**loan_months - 1) / investment_rate_monthly) * (1 + investment_rate_monthly)

# ============================================================================
# SHEET 1: SUMMARY & COMPARISON
# ============================================================================
ws_summary = wb.create_sheet("Summary & Winner", 0)

# Title
ws_summary['A1'] = "6-YEAR INVESTMENT STRATEGY COMPARISON"
ws_summary['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_summary['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_summary.merge_cells('A1:F1')
ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")

# Common Parameters
row = 3
ws_summary[f'A{row}'] = "COMMON PARAMETERS"
ws_summary[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
ws_summary[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws_summary.merge_cells(f'A{row}:B{row}')

row += 1
ws_summary[f'A{row}'] = "Monthly Cash Outflow (Same for both)"
ws_summary[f'B{row}'] = emi
ws_summary[f'B{row}'].number_format = '₹#,##0.00'
ws_summary[f'B{row}'].font = Font(bold=True)

row += 1
ws_summary[f'A{row}'] = "Investment Horizon"
ws_summary[f'B{row}'] = "6 Years"
ws_summary[f'B{row}'].font = Font(bold=True)

row += 1
ws_summary[f'A{row}'] = "Investment Return Rate"
ws_summary[f'B{row}'] = 0.12
ws_summary[f'B{row}'].number_format = '0.00%'
ws_summary[f'B{row}'].font = Font(bold=True)

# Scenario 1
row = 3
ws_summary[f'D{row}'] = "SCENARIO 1: LOAN + LUMPSUM"
ws_summary[f'D{row}'].font = Font(size=12, bold=True, color="FFFFFF")
ws_summary[f'D{row}'].fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
ws_summary.merge_cells(f'D{row}:E{row}')

row += 1
ws_summary[f'D{row}'] = "1. Take Loan"
ws_summary[f'E{row}'] = loan_amount
ws_summary[f'E{row}'].number_format = '₹#,##0'

row += 1
ws_summary[f'D{row}'] = "2. Loan Interest Rate"
ws_summary[f'E{row}'] = loan_rate_annual
ws_summary[f'E{row}'].number_format = '0.00%'

row += 1
ws_summary[f'D{row}'] = "3. Invest Lumpsum @ 12%"
ws_summary[f'E{row}'] = loan_amount
ws_summary[f'E{row}'].number_format = '₹#,##0'

row += 1
ws_summary[f'D{row}'] = "4. Monthly EMI to pay"
ws_summary[f'E{row}'] = emi
ws_summary[f'E{row}'].number_format = '₹#,##0.00'

row += 1
ws_summary[f'D{row}'] = "5. Total EMI Paid (6 years)"
ws_summary[f'E{row}'] = total_emi_paid
ws_summary[f'E{row}'].number_format = '₹#,##0'

row += 1
ws_summary[f'D{row}'] = "6. Investment Value @ 6 years"
ws_summary[f'E{row}'] = lumpsum_value_6y
ws_summary[f'E{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'].font = Font(bold=True, color="C65911")

row += 1
ws_summary[f'D{row}'] = "7. NET VALUE (6 - 5)"
ws_summary[f'E{row}'] = net_value_scenario1
ws_summary[f'E{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'].font = Font(bold=True, size=12, color="C65911")
ws_summary[f'E{row}'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

# Scenario 2
row = 12
ws_summary[f'D{row}'] = "SCENARIO 2: DIRECT SIP"
ws_summary[f'D{row}'].font = Font(size=12, bold=True, color="FFFFFF")
ws_summary[f'D{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws_summary.merge_cells(f'D{row}:E{row}')

row += 1
ws_summary[f'D{row}'] = "1. No Loan taken"
ws_summary[f'E{row}'] = "₹0"

row += 1
ws_summary[f'D{row}'] = "2. Monthly SIP Amount"
ws_summary[f'E{row}'] = emi
ws_summary[f'E{row}'].number_format = '₹#,##0.00'

row += 1
ws_summary[f'D{row}'] = "3. SIP Return Rate"
ws_summary[f'E{row}'] = investment_rate_annual
ws_summary[f'E{row}'].number_format = '0.00%'

row += 1
ws_summary[f'D{row}'] = "4. Total Investment (6 years)"
ws_summary[f'E{row}'] = emi * 72
ws_summary[f'E{row}'].number_format = '₹#,##0'

row += 1
ws_summary[f'D{row}'] = "5. SIP Value @ 6 years"
ws_summary[f'E{row}'] = sip_value_6y
ws_summary[f'E{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'].font = Font(bold=True, size=12, color="4472C4")
ws_summary[f'E{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Comparison
row = 19
ws_summary[f'A{row}'] = "FINAL COMPARISON @ 6 YEARS"
ws_summary[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
ws_summary[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_summary.merge_cells(f'A{row}:F{row}')
ws_summary[f'A{row}'].alignment = Alignment(horizontal="center")

row += 1
headers = ["Metric", "Scenario 1 (Loan)", "Scenario 2 (SIP)", "Difference", "Winner"]
for col, header in enumerate(headers, 1):
    cell = ws_summary.cell(row, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

row += 1
ws_summary[f'A{row}'] = "Monthly Outflow"
ws_summary[f'B{row}'] = emi
ws_summary[f'B{row}'].number_format = '₹#,##0'
ws_summary[f'C{row}'] = emi
ws_summary[f'C{row}'].number_format = '₹#,##0'
ws_summary[f'D{row}'] = 0
ws_summary[f'D{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'] = "EQUAL"
ws_summary[f'E{row}'].font = Font(color="70AD47")

row += 1
ws_summary[f'A{row}'] = "Total Cash Paid"
ws_summary[f'B{row}'] = total_emi_paid
ws_summary[f'B{row}'].number_format = '₹#,##0'
ws_summary[f'C{row}'] = total_emi_paid
ws_summary[f'C{row}'].number_format = '₹#,##0'
ws_summary[f'D{row}'] = 0
ws_summary[f'D{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'] = "EQUAL"
ws_summary[f'E{row}'].font = Font(color="70AD47")

row += 1
ws_summary[f'A{row}'] = "Final Corpus Value"
ws_summary[f'B{row}'] = net_value_scenario1
ws_summary[f'B{row}'].number_format = '₹#,##0'
ws_summary[f'C{row}'] = sip_value_6y
ws_summary[f'C{row}'].number_format = '₹#,##0'
ws_summary[f'D{row}'] = sip_value_6y - net_value_scenario1
ws_summary[f'D{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'] = "Scenario 2" if sip_value_6y > net_value_scenario1 else "Scenario 1"
ws_summary[f'E{row}'].font = Font(bold=True, color="00B050" if sip_value_6y > net_value_scenario1 else "C65911")

row += 1
ws_summary[f'A{row}'] = "Net Gain"
ws_summary[f'B{row}'] = net_value_scenario1 - total_emi_paid
ws_summary[f'B{row}'].number_format = '₹#,##0'
ws_summary[f'C{row}'] = sip_value_6y - total_emi_paid
ws_summary[f'C{row}'].number_format = '₹#,##0'
ws_summary[f'D{row}'] = (sip_value_6y - total_emi_paid) - (net_value_scenario1 - total_emi_paid)
ws_summary[f'D{row}'].number_format = '₹#,##0'
ws_summary[f'E{row}'] = "Scenario 2" if sip_value_6y > net_value_scenario1 else "Scenario 1"
ws_summary[f'E{row}'].font = Font(bold=True, color="00B050" if sip_value_6y > net_value_scenario1 else "C65911")

# Winner announcement
row = 26
ws_summary[f'A{row}'] = "🏆 WINNER"
ws_summary[f'A{row}'].font = Font(size=16, bold=True, color="FFFFFF")
ws_summary[f'A{row}'].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
ws_summary[f'A{row}'].alignment = Alignment(horizontal="center")

winner_text = "SCENARIO 2: DIRECT SIP" if sip_value_6y > net_value_scenario1 else "SCENARIO 1: LOAN + LUMPSUM"
ws_summary[f'B{row}'] = winner_text
ws_summary[f'B{row}'].font = Font(size=16, bold=True, color="00B050")
ws_summary.merge_cells(f'B{row}:F{row}')

row += 1
advantage = abs(sip_value_6y - net_value_scenario1)
ws_summary[f'A{row}'] = "Advantage:"
ws_summary[f'A{row}'].font = Font(size=12, bold=True)
ws_summary[f'B{row}'] = f"₹{advantage:,.0f} higher corpus value!"
ws_summary[f'B{row}'].font = Font(size=12, bold=True, color="00B050")
ws_summary.merge_cells(f'B{row}:F{row}')

# Column widths
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 20
ws_summary.column_dimensions['C'].width = 20
ws_summary.column_dimensions['D'].width = 22
ws_summary.column_dimensions['E'].width = 20
ws_summary.column_dimensions['F'].width = 15

# ============================================================================
# SHEET 2: MONTH-BY-MONTH SCENARIO 1
# ============================================================================
ws_s1 = wb.create_sheet("Scenario 1 - Loan Monthly")

ws_s1['A1'] = "SCENARIO 1: LOAN + LUMPSUM - MONTHLY BREAKDOWN"
ws_s1['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws_s1['A1'].fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
ws_s1.merge_cells('A1:G1')

headers = ["Month", "Loan Outstanding", "EMI Paid", "Principal", "Interest", "Investment Value", "Net Worth"]
for col, header in enumerate(headers, 1):
    cell = ws_s1.cell(2, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

outstanding = loan_amount
cumulative_emi = 0

for month in range(1, 73):
    row = month + 2

    ws_s1[f'A{row}'] = month
    ws_s1[f'B{row}'] = outstanding
    ws_s1[f'B{row}'].number_format = '₹#,##0'

    ws_s1[f'C{row}'] = emi
    ws_s1[f'C{row}'].number_format = '₹#,##0.00'

    # Calculate interest and principal
    interest = outstanding * loan_rate_monthly
    principal = emi - interest

    ws_s1[f'D{row}'] = principal
    ws_s1[f'D{row}'].number_format = '₹#,##0.00'

    ws_s1[f'E{row}'] = interest
    ws_s1[f'E{row}'].number_format = '₹#,##0.00'

    # Investment value (compounded monthly)
    investment_value = loan_amount * (1 + investment_rate_monthly)**month
    ws_s1[f'F{row}'] = investment_value
    ws_s1[f'F{row}'].number_format = '₹#,##0'

    cumulative_emi += emi
    net_worth = investment_value - cumulative_emi
    ws_s1[f'G{row}'] = net_worth
    ws_s1[f'G{row}'].number_format = '₹#,##0'

    if net_worth > 0:
        ws_s1[f'G{row}'].font = Font(color="006100")
    else:
        ws_s1[f'G{row}'].font = Font(color="9C0006")

    outstanding -= principal

# Column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws_s1.column_dimensions[col].width = 18

# ============================================================================
# SHEET 3: MONTH-BY-MONTH SCENARIO 2
# ============================================================================
ws_s2 = wb.create_sheet("Scenario 2 - SIP Monthly")

ws_s2['A1'] = "SCENARIO 2: DIRECT SIP - MONTHLY BREAKDOWN"
ws_s2['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws_s2['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws_s2.merge_cells('A1:E1')

headers = ["Month", "Monthly SIP", "Total Invested", "SIP Value", "Net Gain"]
for col, header in enumerate(headers, 1):
    cell = ws_s2.cell(2, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

for month in range(1, 73):
    row = month + 2

    ws_s2[f'A{row}'] = month

    ws_s2[f'B{row}'] = emi
    ws_s2[f'B{row}'].number_format = '₹#,##0.00'

    total_invested = emi * month
    ws_s2[f'C{row}'] = total_invested
    ws_s2[f'C{row}'].number_format = '₹#,##0'

    # SIP value after n months
    sip_value = emi * (((1 + investment_rate_monthly)**month - 1) / investment_rate_monthly) * (1 + investment_rate_monthly)
    ws_s2[f'D{row}'] = sip_value
    ws_s2[f'D{row}'].number_format = '₹#,##0'

    gain = sip_value - total_invested
    ws_s2[f'E{row}'] = gain
    ws_s2[f'E{row}'].number_format = '₹#,##0'
    ws_s2[f'E{row}'].font = Font(color="006100")

# Column widths
for col in ['A', 'B', 'C', 'D', 'E']:
    ws_s2.column_dimensions[col].width = 18

# ============================================================================
# SHEET 4: YEAR-BY-YEAR COMPARISON
# ============================================================================
ws_yearly = wb.create_sheet("Year-by-Year Comparison")

ws_yearly['A1'] = "YEAR-BY-YEAR COMPARISON"
ws_yearly['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws_yearly['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_yearly.merge_cells('A1:F1')

headers = ["Year", "Scenario 1: Net Worth", "Scenario 2: SIP Value", "Difference", "Better Strategy"]
for col, header in enumerate(headers, 1):
    cell = ws_yearly.cell(2, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

for year in range(1, 7):
    row = year + 2
    month = year * 12

    ws_yearly[f'A{row}'] = year

    # Scenario 1: Lumpsum value - cumulative EMI
    lumpsum_val = loan_amount * (1 + investment_rate_monthly)**month
    cumulative_emi_paid = emi * month
    net_s1 = lumpsum_val - cumulative_emi_paid
    ws_yearly[f'B{row}'] = net_s1
    ws_yearly[f'B{row}'].number_format = '₹#,##0'

    # Scenario 2: SIP value
    sip_val = emi * (((1 + investment_rate_monthly)**month - 1) / investment_rate_monthly) * (1 + investment_rate_monthly)
    ws_yearly[f'C{row}'] = sip_val
    ws_yearly[f'C{row}'].number_format = '₹#,##0'

    # Difference
    diff = sip_val - net_s1
    ws_yearly[f'D{row}'] = diff
    ws_yearly[f'D{row}'].number_format = '₹#,##0'

    # Winner
    ws_yearly[f'E{row}'] = "SIP" if diff > 0 else "Loan+Lumpsum"
    ws_yearly[f'E{row}'].font = Font(bold=True, color="00B050" if diff > 0 else "C65911")

# Column widths
for col in ['A', 'B', 'C', 'D', 'E']:
    ws_yearly.column_dimensions[col].width = 22

# Save workbook
wb.save('/Users/anuragmishra/Code/interview-prep/Investment_Comparison_6Years.xlsx')

print("=" * 80)
print("6-YEAR INVESTMENT STRATEGY COMPARISON")
print("=" * 80)
print(f"\nMonthly Cash Outflow (Same for both): ₹{emi:,.2f}")
print(f"Investment Horizon: 6 Years (72 months)")
print("\n" + "=" * 80)

print(f"\nSCENARIO 1: LOAN + LUMPSUM INVESTMENT")
print(f"  • Take Loan: ₹{loan_amount:,}")
print(f"  • Loan Interest: {loan_rate_annual:.0%} p.a.")
print(f"  • Invest ₹{loan_amount:,} at {investment_rate_annual:.0%} p.a.")
print(f"  • Monthly EMI: ₹{emi:,.2f}")
print(f"  • Total EMI Paid (6 years): ₹{total_emi_paid:,.0f}")
print(f"  • Investment Value @ 6 years: ₹{lumpsum_value_6y:,.0f}")
print(f"  • NET CORPUS: ₹{net_value_scenario1:,.0f}")

print(f"\nSCENARIO 2: DIRECT SIP")
print(f"  • No Loan")
print(f"  • Monthly SIP: ₹{emi:,.2f}")
print(f"  • SIP Returns: {investment_rate_annual:.0%} p.a.")
print(f"  • Total Invested (6 years): ₹{total_emi_paid:,.0f}")
print(f"  • SIP CORPUS: ₹{sip_value_6y:,.0f}")

print("\n" + "=" * 80)
print("FINAL COMPARISON @ 6 YEARS")
print("=" * 80)

if sip_value_6y > net_value_scenario1:
    winner = "SCENARIO 2: DIRECT SIP"
    advantage = sip_value_6y - net_value_scenario1
    print(f"\n🏆 WINNER: {winner}")
    print(f"   Advantage: ₹{advantage:,.0f} higher corpus!")
else:
    winner = "SCENARIO 1: LOAN + LUMPSUM"
    advantage = net_value_scenario1 - sip_value_6y
    print(f"\n🏆 WINNER: {winner}")
    print(f"   Advantage: ₹{advantage:,.0f} higher corpus!")

print("\n" + "=" * 80)
print("Excel file created: Investment_Comparison_6Years.xlsx")
print("=" * 80)
